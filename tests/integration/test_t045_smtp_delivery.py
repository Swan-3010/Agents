"""T-045: Real SMTP delivery integration tests

Goal:
- Send test report via real SMTP to sandbox inbox
- Verify email delivery, attachment integrity, subject match

AC:
- Email delivered successfully
- Attachment not corrupted (valid XLSX)
- Subject matches expected pattern
"""

import os
import pytest
import smtplib
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from pathlib import Path
import openpyxl
import time
import imaplib
import email
from email.header import decode_header


class TestSMTPDelivery:
    """Real SMTP delivery to sandbox inbox"""

    @pytest.fixture
    def smtp_config(self):
        """SMTP configuration from environment"""
        return {
            "host": os.getenv("SMTP_HOST", "smtp.yandex.ru"),
            "port": int(os.getenv("SMTP_PORT", "465")),
            "user": os.getenv("SMTP_USER"),
            "password": os.getenv("SMTP_PASSWORD"),
            "from_addr": os.getenv("SMTP_FROM"),
            "to_addr": os.getenv("SMTP_TO_SANDBOX"),
        }

    @pytest.fixture
    def imap_config(self):
        """IMAP configuration for verification"""
        return {
            "host": os.getenv("IMAP_HOST", "imap.yandex.ru"),
            "port": int(os.getenv("IMAP_PORT", "993")),
            "user": os.getenv("IMAP_USER"),
            "password": os.getenv("IMAP_PASSWORD"),
        }

    @pytest.fixture
    def test_xlsx_file(self):
        """Create a valid test XLSX file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Report"
            ws["A1"] = "Receipt ID"
            ws["B1"] = "Amount"
            ws["C1"] = "Date"
            ws["A2"] = "RCP-001"
            ws["B2"] = 150.50
            ws["C2"] = "2025-01-15"
            wb.save(tmp.name)
            yield tmp.name
        os.unlink(tmp.name)

    def test_smtp_send_with_attachment(self, smtp_config, test_xlsx_file):
        """AC: Send email with XLSX attachment via SMTP"""
        # Skip if no credentials
        if not smtp_config["user"] or not smtp_config["password"]:
            pytest.skip("SMTP credentials not configured")

        # Create message
        msg = MIMEMultipart()
        msg["From"] = smtp_config["from_addr"] or smtp_config["user"]
        msg["To"] = smtp_config["to_addr"] or smtp_config["user"]
        msg["Subject"] = "[Test] Receipt Report T-045"
        
        # Add body
        body = "Test email for T-045 SMTP delivery integration test."
        msg.attach(MIMEText(body, "plain"))
        
        # Attach XLSX file
        with open(test_xlsx_file, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename=receipt_report.xlsx",
        )
        msg.attach(part)
        
        # Send via SMTP
        with smtplib.SMTP_SSL(smtp_config["host"], smtp_config["port"]) as server:
            server.login(smtp_config["user"], smtp_config["password"])
            server.send_message(msg)
        
        # If we got here, email was sent successfully
        assert True, "Email sent successfully"

    def test_verify_email_delivery(self, smtp_config, imap_config, test_xlsx_file):
        """AC: Verify email delivery, subject match, attachment integrity"""
        # Skip if no credentials
        if not smtp_config["user"] or not imap_config["user"]:
            pytest.skip("SMTP/IMAP credentials not configured")

        subject = "[Test] Receipt Report T-045"
        
        # Send email first
        msg = MIMEMultipart()
        msg["From"] = smtp_config["from_addr"] or smtp_config["user"]
        msg["To"] = smtp_config["to_addr"] or smtp_config["user"]
        msg["Subject"] = subject
        msg.attach(MIMEText("Verification test for T-045", "plain"))
        
        with open(test_xlsx_file, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment; filename=report.xlsx")
        msg.attach(part)
        
        with smtplib.SMTP_SSL(smtp_config["host"], smtp_config["port"]) as server:
            server.login(smtp_config["user"], smtp_config["password"])
            server.send_message(msg)
        
        # Wait for delivery
        time.sleep(5)
        
        # Check IMAP inbox
        with imaplib.IMAP4_SSL(imap_config["host"], imap_config["port"]) as mail:
            mail.login(imap_config["user"], imap_config["password"])
            mail.select("INBOX")
            
            # Search for the email
            _, data = mail.search(None, f'SUBJECT "{subject}"')
            mail_ids = data[0].split()
            
            # AC: Email delivered
            assert len(mail_ids) > 0, "Email not found in inbox"
            
            # Fetch latest matching email
            _, msg_data = mail.fetch(mail_ids[-1], "(RFC822)")
            email_body = msg_data[0][1]
            email_msg = email.message_from_bytes(email_body)
            
            # AC: Subject matches
            email_subject = decode_header(email_msg["Subject"])[0][0]
            if isinstance(email_subject, bytes):
                email_subject = email_subject.decode()
            assert subject in email_subject, f"Subject mismatch: {email_subject}"
            
            # AC: Attachment not corrupted
            attachment_found = False
            for part in email_msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue
                if part.get("Content-Disposition") is None:
                    continue
                
                filename = part.get_filename()
                if filename and filename.endswith(".xlsx"):
                    attachment_found = True
                    # Save attachment temporarily
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                        tmp.write(part.get_payload(decode=True))
                        tmp_path = tmp.name
                    
                    try:
                        # Try to open XLSX to verify integrity
                        wb = openpyxl.load_workbook(tmp_path)
                        assert wb is not None, "XLSX file corrupted"
                        wb.close()
                    finally:
                        os.unlink(tmp_path)
            
            assert attachment_found, "XLSX attachment not found in email"
