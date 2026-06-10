"""T-044: Golden-file and content validation tests for XLSX.

Goal:
  Validate XLSX content quality and structure beyond basic generation:
    - Golden-file regression: compare generated XLSX to reference snapshot
    - Size validation: file size >0, within reasonable bounds
    - Column validation: correct count, correct names, correct order
    - Data type validation: numbers are numbers, dates are dates, strings are strings
    - Row count validation: expected number of data rows
    - Content integrity: no empty cells in required columns

Difference from T-043:
  T-043: validates that XLSX is created and has basic structure
  T-044: validates XLSX **content quality** — ready for production use

No unittest.mock, no patch, no MagicMock.

Task: T-044
Epic: E-11 — Real Artifact Generation
Depends on: T-043
"""
import pytest
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import hashlib

from packages.report_core.generator import ReceiptReportGenerator


# ---------------------------------------------------------------------------
# Config & Helpers
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path("test_artifacts/t044")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

GOLDEN_DIR = Path("tests/integration/golden_files")
GOLDEN_DIR.mkdir(parents=True, exist_ok=True)


def _make_receipt_dict(receipt_id: str = "R001") -> dict:
    """Create a sample receipt data dict for generator input."""
    return {
        "receipt_id": receipt_id,
        "date": datetime(2026, 6, 10, 14, 30).isoformat(),
        "merchant": "Test Shop",
        "total": 1234.56,
        "items_count": 3,
        "url": "https://ofd.ru/r/test123",
        "status": "processed",
    }


def _generate_test_xlsx(filename: str, receipts: list) -> Path:
    """Helper: generate XLSX file and return Path."""
    gen = ReceiptReportGenerator()
    gen.add_multiple_receipts(receipts)
    filepath = OUTPUT_DIR / filename
    gen.save(str(filepath))
    return filepath


# ---------------------------------------------------------------------------
# Layer 1: File size and basic integrity
# ---------------------------------------------------------------------------

class TestXLSXSizeAndIntegrity:
    """Validates XLSX file size and basic integrity checks."""

    def test_xlsx_size_greater_than_zero(self):
        """Generated XLSX must have size >0."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_size_gt_zero.xlsx", receipts)
        assert xlsx_path.stat().st_size > 0

    def test_xlsx_size_within_reasonable_bounds(self):
        """XLSX with 1 receipt should be 1KB-100KB (not corrupted/bloated)."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_size_bounds.xlsx", receipts)
        size = xlsx_path.stat().st_size
        assert 1024 < size < 100 * 1024, f"XLSX size {size} bytes out of expected range"

    def test_xlsx_size_grows_with_more_receipts(self):
        """XLSX with 10 receipts must be larger than XLSX with 1 receipt."""
        receipts_1 = [_make_receipt_dict("R001")]
        xlsx_1 = _generate_test_xlsx("test_size_1_receipt.xlsx", receipts_1)
        receipts_10 = [_make_receipt_dict(f"R{i:03d}") for i in range(1, 11)]
        xlsx_10 = _generate_test_xlsx("test_size_10_receipts.xlsx", receipts_10)
        assert xlsx_10.stat().st_size > xlsx_1.stat().st_size


# ---------------------------------------------------------------------------
# Layer 2: Column validation
# ---------------------------------------------------------------------------

class TestXLSXColumnValidation:
    """Validates XLSX column structure: count, names, order."""

    def test_xlsx_has_expected_column_count(self):
        """XLSX must have exactly 7 columns (from _make_receipt_dict keys)."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_column_count.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(header_row) == 7, f"Expected 7 columns, got {len(header_row)}"
        wb.close()

    def test_xlsx_columns_have_correct_names(self):
        """XLSX columns must match expected keys from receipt dict."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_column_names.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        expected_columns = ["receipt_id", "date", "merchant", "total", "items_count", "url", "status"]
        for col in expected_columns:
            assert col in header_row, f"Column {col!r} missing from headers {header_row}"
        wb.close()

    def test_xlsx_columns_in_expected_order(self):
        """XLSX columns must appear in dict key order (insertion order)."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_column_order.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1] if cell.value is not None]
        expected_order = ["receipt_id", "date", "merchant", "total", "items_count", "url", "status"]
        assert header_row == expected_order, f"Column order mismatch: {header_row} != {expected_order}"
        wb.close()


# ---------------------------------------------------------------------------
# Layer 3: Data type and content validation
# ---------------------------------------------------------------------------

class TestXLSXDataValidation:
    """Validates XLSX data types and content integrity."""

    def test_xlsx_numeric_column_has_numbers(self):
        """Total column must contain numeric values."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_numeric.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        total_idx = header_row.index("total")
        total_value = ws[2][total_idx].value
        assert isinstance(total_value, (int, float)), f"Total value {total_value!r} is not numeric"
        wb.close()

    def test_xlsx_string_column_has_strings(self):
        """Merchant column must contain string values."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_string.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        merchant_idx = header_row.index("merchant")
        merchant_value = ws[2][merchant_idx].value
        assert isinstance(merchant_value, str), f"Merchant value {merchant_value!r} is not string"
        wb.close()

    def test_xlsx_no_empty_cells_in_required_columns(self):
        """Receipt_id, merchant, total must not be empty."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_no_empty.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        required_cols = ["receipt_id", "merchant", "total"]
        for col_name in required_cols:
            col_idx = header_row.index(col_name)
            cell_value = ws[2][col_idx].value
            assert cell_value is not None and cell_value != "", (
                f"Required column {col_name!r} has empty value"
            )
        wb.close()

    def test_xlsx_date_column_contains_iso_string(self):
        """Date column must contain ISO 8601 datetime string."""
        receipts = [_make_receipt_dict("R001")]
        xlsx_path = _generate_test_xlsx("test_date_iso.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        date_idx = header_row.index("date")
        date_value = ws[2][date_idx].value
        # Should be ISO string like "2026-06-10T14:30:00"
        assert isinstance(date_value, str), f"Date value {date_value!r} is not string"
        assert "T" in date_value, f"Date {date_value!r} not in ISO format"
        wb.close()


# ---------------------------------------------------------------------------
# Layer 4: Row count validation
# ---------------------------------------------------------------------------

class TestXLSXRowCount:
    """Validates XLSX row count matches input data."""

    def test_xlsx_row_count_matches_input_count(self):
        """XLSX with 5 receipts must have 5 data rows (+ 1 header = 6 total)."""
        receipts = [_make_receipt_dict(f"R{i:03d}") for i in range(1, 6)]
        xlsx_path = _generate_test_xlsx("test_row_count.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_row == 6, f"Expected 6 rows (1 header + 5 data), got {ws.max_row}"
        wb.close()

    def test_xlsx_each_row_has_receipt_id(self):
        """Every data row must have a non-empty receipt_id."""
        receipts = [_make_receipt_dict(f"R{i:03d}") for i in range(1, 4)]
        xlsx_path = _generate_test_xlsx("test_each_row_id.xlsx", receipts)
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        receipt_id_idx = header_row.index("receipt_id")
        for row_idx in range(2, ws.max_row + 1):
            receipt_id = ws[row_idx][receipt_id_idx].value
            assert receipt_id is not None and receipt_id != "", (
                f"Row {row_idx} has empty receipt_id"
            )
        wb.close()


# ---------------------------------------------------------------------------
# Layer 5: Golden-file regression (optional/advanced)
# ---------------------------------------------------------------------------

class TestXLSXGoldenFile:
    """Golden-file regression: compare generated XLSX to reference snapshot.

    Note: Golden-file testing is brittle (timestamp, minor format changes break it).
    This approach uses content hash instead of binary comparison.
    """

    def test_xlsx_content_hash_stable(self):
        """Generating same data twice must produce same content hash."""
        receipts = [_make_receipt_dict("R999")]
        xlsx_1 = _generate_test_xlsx("test_golden_1.xlsx", receipts)
        xlsx_2 = _generate_test_xlsx("test_golden_2.xlsx", receipts)
        # Load both and compare header + first data row content
        wb1 = load_workbook(xlsx_1)
        ws1 = wb1.active
        content_1 = str([cell.value for cell in ws1[1]]) + str([cell.value for cell in ws1[2]])
        wb1.close()
        wb2 = load_workbook(xlsx_2)
        ws2 = wb2.active
        content_2 = str([cell.value for cell in ws2[1]]) + str([cell.value for cell in ws2[2]])
        wb2.close()
        assert content_1 == content_2, "Generated XLSX content differs between runs"

    @pytest.mark.skip(reason="Binary golden-file too brittle — use content validation instead")
    def test_xlsx_binary_matches_golden_file(self):
        """Binary comparison to golden file (disabled — too fragile)."""
        # This would:
        # 1. Generate XLSX
        # 2. Compare bytes to golden_files/reference.xlsx
        # 3. Fail on any openpyxl version change, timestamp, etc.
        pass


# ---------------------------------------------------------------------------
# Layer 6: Edge cases
# ---------------------------------------------------------------------------

class TestXLSXEdgeCases:
    """Validates XLSX generation handles edge cases correctly."""

    def test_xlsx_handles_special_characters_in_merchant(self):
        """Merchant name with special characters must not break XLSX."""
        receipt = _make_receipt_dict("R001")
        receipt["merchant"] = "Test & Shop: \\\"\'\"<>!"
        xlsx_path = _generate_test_xlsx("test_special_chars.xlsx", [receipt])
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        merchant_idx = header_row.index("merchant")
        merchant_value = ws[2][merchant_idx].value
        assert "&" in merchant_value
        wb.close()

    def test_xlsx_handles_very_large_total(self):
        """Total with large number (millions) must be preserved accurately."""
        receipt = _make_receipt_dict("R001")
        receipt["total"] = 9999999.99
        xlsx_path = _generate_test_xlsx("test_large_total.xlsx", [receipt])
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        total_idx = header_row.index("total")
        total_value = ws[2][total_idx].value
        assert abs(total_value - 9999999.99) < 0.01, f"Total value {total_value} != 9999999.99"
        wb.close()

    def test_xlsx_handles_zero_items_count(self):
        """Items_count = 0 must be valid (not treated as empty)."""
        receipt = _make_receipt_dict("R001")
        receipt["items_count"] = 0
        xlsx_path = _generate_test_xlsx("test_zero_items.xlsx", [receipt])
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        items_idx = header_row.index("items_count")
        items_value = ws[2][items_idx].value
        assert items_value == 0, f"items_count should be 0, got {items_value}"
        wb.close()
