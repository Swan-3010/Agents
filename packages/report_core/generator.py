"""Модуль report_core.generator - генератор XLSX отчетов о чеках.

Используется для выполнения задачи T-029:
- Создание XLSX файла с данными чека
- Форматирование таблицы с заголовками и данными
- Сохранение в файл или в память
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any, Optional
from datetime import datetime
import io
import logging

logger = logging.getLogger(__name__)


class ReceiptReportGenerator:
    """Генератор XLSX отчетов о чеках.
    
    Attributes:
        workbook: Экземпляр Workbook для создания XLSX
        worksheet: Активный лист в workbook
    """
    
    def __init__(self, sheet_name: str = "Чеки"):
        """Инициализация генератора.
        
        Args:
            sheet_name: Название листа в XLSX (по умолчанию "Чеки")
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        logger.info(f"Создан новый workbook с листом '{sheet_name}'")
    
    def _setup_headers(self, headers: List[str]) -> None:
        """Настроить заголовки таблицы.
        
        Args:
            headers: Список заголовков колонок
        """
        # Стиль заголовка
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Запись заголовков
        for col_idx, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        logger.info(f"Заголовки установлены: {headers}")
    
    def add_receipt_data(self, receipt_data: Dict[str, Any]) -> int:
        """Добавить данные одного чека в таблицу.
        
        Args:
            receipt_data: Словарь с данными чека
            
        Returns:
            Номер строки, куда были добавлены данные
        """
        # Если это первая запись, создаем заголовки
        if self.worksheet.max_row == 1 and self.worksheet.cell(1, 1).value is None:
            headers = list(receipt_data.keys())
            self._setup_headers(headers)
        
        # Номер новой строки
        row_idx = self.worksheet.max_row + 1
        
        # Запись данных
        for col_idx, value in enumerate(receipt_data.values(), start=1):
            cell = self.worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        logger.info(f"Добавлена строка {row_idx} с данными чека")
        return row_idx
    
    def add_multiple_receipts(self, receipts: List[Dict[str, Any]]) -> int:
        """Добавить несколько чеков в таблицу.
        
        Args:
            receipts: Список словарей с данными чеков
            
        Returns:
            Количество добавленных строк
        """
        count = 0
        for receipt in receipts:
            self.add_receipt_data(receipt)
            count += 1
        
        logger.info(f"Добавлено {count} чеков в таблицу")
        return count
    
    def auto_adjust_columns(self) -> None:
        """Автоматическая подстройка ширины колонок."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Ширина колонок подстроена")
    
    def save(self, filepath: str) -> None:
        """Сохранить XLSX файл.
        
        Args:
            filepath: Путь к файлу для сохранения
        """
        self.auto_adjust_columns()
        self.workbook.save(filepath)
        logger.info(f"Отчет сохранен в файл: {filepath}")
    
    def save_to_bytes(self) -> bytes:
        """Сохранить XLSX в память (в виде bytes).
        
        Returns:
            XLSX файл в виде bytes
        """
        self.auto_adjust_columns()
        virtual_file = io.BytesIO()
        self.workbook.save(virtual_file)
        virtual_file.seek(0)
        logger.info("Отчет сохранен в память")
        return virtual_file.read()


def create_receipt_report(
    receipts: List[Dict[str, Any]],
    filepath: Optional[str] = None,
    sheet_name: str = "Чеки"
) -> Optional[bytes]:
    """Вспомогательная функция для создания отчета о чеках.
    
    Args:
        receipts: Список чеков для добавления в отчет
        filepath: Путь к файлу (если None, возвращает bytes)
        sheet_name: Название листа
    
    Returns:
        bytes если filepath is None, иначе None
    """
    generator = ReceiptReportGenerator(sheet_name=sheet_name)
    generator.add_multiple_receipts(receipts)
    
    if filepath:
        generator.save(filepath)
        return None
    else:
        return generator.save_to_bytes()


def create_single_receipt_report(
    receipt_data: Dict[str, Any],
    filepath: Optional[str] = None
) -> Optional[bytes]:
    """Создать отчет для одного чека.
    
    Args:
        receipt_data: Данные одного чека
        filepath: Путь к файлу (если None, возвращает bytes)
    
    Returns:
        bytes если filepath is None, иначе None
    """
    return create_receipt_report([receipt_data], filepath=filepath)
