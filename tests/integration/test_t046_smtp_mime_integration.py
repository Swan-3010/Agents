"""T-046: SMTP MIME attachment integration tests

Goal:
- Test MIME email composition with XLSX attachment
- Verify attachment metadata (filename, size, content-type)
- Verify email headers (subject, from, to, body)
- Happy-path test report delivery

AC:
- Attachment correctly encoded as MIME part
- Filename preserved in Content-Disposition header
- Subject, body match expected pattern
- Email structure valid for delivery
"""

import os
import pytest
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from email import message_from_string
import openpyxl
import base64


class TestSMTPMIMEIntegration:
    """MIME attachment and email composition tests"""

    @pytest.fixture
    def test_xlsx_file(self):
        """Create test XLSX file"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Receipt Report"
            ws["A1"] = "ID"
            ws["B1"] = "Amount"
            ws["A2"] = "R001"
            ws["B2"] = 100.50
            wb.save(tmp.name)
            yield tmp.name
        os.unlink(tmp.name)

    def test_mime_message_structure(self, test_xlsx_file):
        """AC: MIME message has correct multipart structure"""
        msg = MIMEMultipart()
        msg["From"] = "sender@example.com"
        msg["To"] = "recipient@example.com"
        msg["Subject"] = "Test Receipt Report"
        
        # Add text body
        body = "Please find attached receipt report."
        msg.attach(MIMEText(body, "plain"))
        
        # Add XLSX attachment
        with open(test_xlsx_file, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment; filename=receipt_report.xlsx")
        msg.attach(part)
        
        # Verify message structure
        msg_str = msg.as_string()
        assert "multipart/mixed" in msg_str, "Message should be multipart/mixed"
        assert "From: sender@example.com" in msg_str
        assert "To: recipient@example.com" in msg_str
        assert "Subject: Test Receipt Report" in msg_str
        assert "receipt_report.xlsx" in msg_str

    def test_attachment_metadata(self, test_xlsx_file):
        """AC: Attachment has correct filename and content-type"""
        msg = MIMEMultipart()
        
        filename = "test_report.xlsx"
        with open(test_xlsx_file, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={filename}",
        )
        msg.attach(part)
        
        # Parse message to verify attachment
        msg_str = msg.as_string()
        parsed_msg = message_from_string(msg_str)
        
        attachment_found = False
        for part in parsed_msg.walk():
            if part.get_content_maintype() == "multipart":
                continue
            if part.get("Content-Disposition") is None:
                continue
            
            fname = part.get_filename()
            if fname:
                attachment_found = True
                # AC: Filename preserved
                assert fname == filename, f"Filename mismatch: {fname}"
                # AC: Content-Type correct
                assert "application" in part.get_content_type()
        
        assert attachment_found, "Attachment not found in message"

    def test_attachment_encoding(self, test_xlsx_file):
        """AC: Attachment is base64 encoded and can be decoded"""
        msg = MIMEMultipart()
        
        # Read original file
        with open(test_xlsx_file, "rb") as f:
            original_content = f.read()
        
        # Attach file
        with open(test_xlsx_file, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment; filename=report.xlsx")
        msg.attach(part)
        
        # Parse and extract attachment
        msg_str = msg.as_string()
        parsed_msg = message_from_string(msg_str)
        
        for part in parsed_msg.walk():
            if part.get_filename() == "report.xlsx":
                # AC: Decode attachment
                decoded_content = part.get_payload(decode=True)
                # AC: Content matches original
                assert decoded_content == original_content, "Decoded content doesn't match original"
                # AC: Content is valid XLSX
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(decoded_content)
                    tmp_path = tmp.name
                try:
                    wb = openpyxl.load_workbook(tmp_path)
                    assert wb is not None
                    wb.close()
                finally:
                    os.unlink(tmp_path)

    def test_email_headers(self):
        """AC: Email has correct headers for delivery"""
        msg = MIMEMultipart()
        msg["From"] = "test@example.com"
        msg["To"] = "recipient@example.com"
        msg["Subject"] = "[Receipt] Daily Report 2025-01-15"
        msg["Date"] = "Wed, 15 Jan 2025 12:00:00 +0000"
        
        body = "Automated receipt report for 2025-01-15."
        msg.attach(MIMEText(body, "plain"))
        
        msg_str = msg.as_string()
        
        # AC: Required headers present
        assert "From: test@example.com" in msg_str
        assert "To: recipient@example.com" in msg_str
        assert "Subject: [Receipt] Daily Report 2025-01-15" in msg_str
        assert "Date:" in msg_str
        
        # AC: Body present
        assert "Automated receipt report" in msg_str

    def test_happy_path_report_composition(self, test_xlsx_file):
        """AC: Complete happy-path test for report email composition"""
        # Compose complete email
        msg = MIMEMultipart()
        msg["From"] = "noreply@receipts.example.com"
        msg["To"] = "reports@example.com"
        msg["Subject"] = "Receipt Processing Report - 2025-01-15"
        
        body_text = (
            "Receipt processing completed successfully.\n\n"
            "Total receipts processed: 5\n"
            "Successful: 4\n"
            "Failed: 1\n\n"
            "Please see attached report for details."
        )
        msg.attach(MIMEText(body_text, "plain"))
        
        # Attach report
        with open(test_xlsx_file, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            "attachment; filename=receipt_report_2025-01-15.xlsx",
        )
        msg.attach(part)
        
        # Verify complete message
        msg_str = msg.as_string()
        
        # AC: All headers correct
        assert "From: noreply@receipts.example.com" in msg_str
        assert "To: reports@example.com" in msg_str
        assert "Subject: Receipt Processing Report - 2025-01-15" in msg_str
        
        # AC: Body contains expected information
        assert "Receipt processing completed" in msg_str
        assert "Total receipts processed: 5" in msg_str
        
        # AC: Attachment present with correct filename
        assert "receipt_report_2025-01-15.xlsx" in msg_str
        
        # AC: Message can be parsed
        parsed_msg = message_from_string(msg_str)
        assert parsed_msg is not None
        
        # AC: Attachment can be extracted and is valid
        attachment_valid = False
        for part in parsed_msg.walk():
            fname = part.get_filename()
            if fname and "receipt_report" in fname:
                attachment_valid = True
                decoded = part.get_payload(decode=True)
                assert len(decoded) > 0, "Attachment is empty"
        
        assert attachment_valid, "Valid attachment not found"
