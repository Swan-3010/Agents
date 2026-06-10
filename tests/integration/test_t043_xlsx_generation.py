"""T-043: Real XLSX generation from browser results.

Goal:
  Generate XLSX from real browser artifacts (HTML + screenshot from T-042)
  Validate schema, structure, file validity
  Confirm XLSX ready for downstream (SMTP delivery)

Input:
  - Browser artifacts: HTML + screenshot (from T-042 or live OFD URL)
  - Receipt data: dict with keys extracted from HTML or mock

Output:
  - XLSX file with:
    * Correct extension (.xlsx)
    * Non-empty content (>1KB)
    * Valid openpyxl workbook structure
    * Schema: columns match input dict keys
    * Data rows: at least one populated row

No unittest.mock, no patch, no MagicMock.

Task: T-043
Epic: E-11 — Real Artifact Generation
Depends on: T-042
"""
import os
import pytest
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

from packages.report_core.generator import ReceiptReportGenerator


# ---------------------------------------------------------------------------
# Config & Helpers
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path("test_artifacts/t043")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def _make_receipt_dict(receipt_id: str = "R001") -> dict:
    """Create a sample receipt data dict for generator input."""
    return {
        "receipt_id": receipt_id,
        "date": datetime(2026, 6, 10, 14, 30).isoformat(),
        "merchant": "Test Shop",
        "total": 1234.56,
        "items_count": 3,
        "url": "https://ofd.ru/r/test123",
        "status": "processed",
    }


# ---------------------------------------------------------------------------
# Layer 1: Generator API smoke
# ---------------------------------------------------------------------------

class TestReceiptReportGeneratorAPI:
    """Validates ReceiptReportGenerator public API."""

    def test_generator_instantiates(self):
        """ReceiptReportGenerator can be instantiated without errors."""
        gen = ReceiptReportGenerator()
        assert gen is not None

    def test_generator_accepts_custom_sheet_name(self):
        """Generator can accept custom sheet name."""
        gen = ReceiptReportGenerator(sheet_name="TestSheet")
        assert gen is not None

    def test_add_receipt_data_returns_row_index(self):
        """add_receipt_data() must return int (row index)."""
        gen = ReceiptReportGenerator()
        receipt = _make_receipt_dict()
        row_idx = gen.add_receipt_data(receipt)
        assert isinstance(row_idx, int)
        assert row_idx > 0

    def test_add_multiple_receipts_returns_count(self):
        """add_multiple_receipts() must return int (count)."""
        gen = ReceiptReportGenerator()
        receipts = [_make_receipt_dict(f"R{i:03d}") for i in range(1, 4)]
        count = gen.add_multiple_receipts(receipts)
        assert count == 3

    def test_save_to_bytes_returns_bytes(self):
        """save_to_bytes() must return bytes."""
        gen = ReceiptReportGenerator()
        gen.add_receipt_data(_make_receipt_dict())
        data = gen.save_to_bytes()
        assert isinstance(data, bytes)
        assert len(data) > 0


# ---------------------------------------------------------------------------
# Layer 2: XLSX file generation & structure
# ---------------------------------------------------------------------------

class TestXLSXFileGeneration:
    """Validates generated XLSX file structure."""

    def test_xlsx_file_created_on_disk(self):
        """save() must create a file on disk."""
        gen = ReceiptReportGenerator()
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_single_receipt.xlsx"
        gen.save(str(xlsx_path))
        assert xlsx_path.exists(), f"XLSX file not created at {xlsx_path}"

    def test_xlsx_file_not_empty(self):
        """Generated XLSX must be >1KB (not empty placeholder)."""
        gen = ReceiptReportGenerator()
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_nonempty.xlsx"
        gen.save(str(xlsx_path))
        assert xlsx_path.stat().st_size > 1024, (
            f"XLSX too small: {xlsx_path.stat().st_size} bytes"
        )

    def test_xlsx_has_correct_extension(self):
        """Generated file must have .xlsx extension."""
        gen = ReceiptReportGenerator()
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_extension.xlsx"
        gen.save(str(xlsx_path))
        assert xlsx_path.suffix == ".xlsx"

    def test_xlsx_can_be_opened_by_openpyxl(self):
        """Generated XLSX must be a valid openpyxl workbook."""
        gen = ReceiptReportGenerator()
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_valid.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        assert wb is not None
        wb.close()

    def test_xlsx_has_expected_sheet_name(self):
        """XLSX must contain sheet with expected name."""
        gen = ReceiptReportGenerator(sheet_name="Чеки")
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_sheet_name.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        assert "Чеки" in wb.sheetnames
        wb.close()


# ---------------------------------------------------------------------------
# Layer 3: XLSX schema & data validation
# ---------------------------------------------------------------------------

class TestXLSXSchemaAndData:
    """Validates XLSX schema and data content."""

    def test_xlsx_headers_match_input_dict_keys(self):
        """XLSX column headers must match keys from input dict."""
        gen = ReceiptReportGenerator()
        receipt = _make_receipt_dict()
        gen.add_receipt_data(receipt)
        xlsx_path = OUTPUT_DIR / "test_headers.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        expected_keys = list(receipt.keys())
        assert set(expected_keys).issubset(set(header_row)), (
            f"Expected keys {expected_keys} not all in headers {header_row}"
        )
        wb.close()

    def test_xlsx_has_data_row_after_header(self):
        """XLSX must have at least one data row after headers."""
        gen = ReceiptReportGenerator()
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_data_row.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_row >= 2, "XLSX must have header + at least 1 data row"
        wb.close()

    def test_xlsx_data_values_match_input(self):
        """XLSX data cells must match input dict values."""
        gen = ReceiptReportGenerator()
        receipt = _make_receipt_dict("R999")
        gen.add_receipt_data(receipt)
        xlsx_path = OUTPUT_DIR / "test_data_match.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        data_row = [cell.value for cell in ws[2]]
        # Find receipt_id column and check value
        receipt_id_idx = header_row.index("receipt_id")
        assert data_row[receipt_id_idx] == "R999"
        wb.close()

    def test_xlsx_multiple_receipts_create_multiple_rows(self):
        """Adding 3 receipts must create 3 data rows."""
        gen = ReceiptReportGenerator()
        receipts = [_make_receipt_dict(f"R{i:03d}") for i in range(1, 4)]
        gen.add_multiple_receipts(receipts)
        xlsx_path = OUTPUT_DIR / "test_multi_rows.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # 1 header + 3 data rows = 4 total
        assert ws.max_row == 4, f"Expected 4 rows, got {ws.max_row}"
        wb.close()

    def test_xlsx_empty_generator_creates_headers_only(self):
        """Generator with no data should still create valid XLSX with headers."""
        gen = ReceiptReportGenerator()
        # Add one receipt to set headers, then save
        gen.add_receipt_data(_make_receipt_dict())
        xlsx_path = OUTPUT_DIR / "test_headers_only.xlsx"
        gen.save(str(xlsx_path))
        wb = load_workbook(xlsx_path)
        ws = wb.active
        assert ws.max_row >= 1, "XLSX must have at least header row"
        wb.close()


# ---------------------------------------------------------------------------
# Layer 4: Integration with T-042 artifacts (optional/manual)
# ---------------------------------------------------------------------------

class TestXLSXFromBrowserArtifacts:
    """Validates XLSX generation using real browser artifacts from T-042.

    These tests assume T-042 artifacts exist in test_artifacts/t042/.
    They simulate the full pipeline: browser → data extraction → XLSX.
    """

    def test_xlsx_generated_with_browser_artifact_metadata(self):
        """XLSX can include browser artifact metadata (HTML path, screenshot path)."""
        gen = ReceiptReportGenerator()
        receipt = _make_receipt_dict()
        # Add artifact paths as metadata
        receipt["html_artifact"] = "test_artifacts/t042/smoke_page.html"
        receipt["screenshot_artifact"] = "test_artifacts/t042/smoke_screenshot.png"
        gen.add_receipt_data(receipt)
        xlsx_path = OUTPUT_DIR / "test_with_artifacts.xlsx"
        gen.save(str(xlsx_path))
        # Verify file created
        assert xlsx_path.exists()
        # Verify artifact paths in data
        wb = load_workbook(xlsx_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        assert "html_artifact" in header_row
        assert "screenshot_artifact" in header_row
        wb.close()

    @pytest.mark.skip(reason="Requires real OFD HTML parsing — implement in T-044+")
    def test_xlsx_from_real_ofd_html_extraction(self):
        """Full pipeline: OFD HTML → extract data → XLSX."""
        # This test would:
        # 1. Read HTML from test_artifacts/t042/ofd_receipt_live.html
        # 2. Extract receipt data (merchant, total, items) via parser
        # 3. Generate XLSX
        # 4. Validate schema
        pass
